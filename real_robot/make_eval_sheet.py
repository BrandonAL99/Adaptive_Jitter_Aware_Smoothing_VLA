"""Generate results/eval_recording_sheet.xlsx — run once to create the sheet."""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
OUT  = ROOT / "results" / "eval_recording_sheet.xlsx"

wb = openpyxl.Workbook()

# ── colour palette ─────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark blue
C_BASELINE = "D9E1F2"   # light blue
C_LP       = "E2EFDA"   # light green
C_AJAS     = "FCE4D6"   # light orange
C_SCORE    = "FFF2CC"   # light yellow
C_GUIDE    = "F2F2F2"   # light grey

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font():
    return Font(bold=True, color="FFFFFF")

def _cell(ws, row, col, value, fill=None, bold=False, align="center", wrap=False, color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _border()
    if fill:
        c.fill = _fill(fill)
    if bold or fill == C_HEADER:
        c.font = Font(bold=True, color=color or ("FFFFFF" if fill == C_HEADER else "000000"))
    elif color:
        c.font = Font(color=color)
    return c


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Scoring Guide
# ══════════════════════════════════════════════════════════════════════════════
ws_guide = wb.active
ws_guide.title = "Scoring Guide"
ws_guide.column_dimensions["A"].width = 22
ws_guide.column_dimensions["B"].width = 55
ws_guide.column_dimensions["C"].width = 14

guide_rows = [
    ("TASK", "CRITERION", "POINTS"),
    ("Pick-place", "Robot grasps the red cube (lifts off table)", 0.5),
    ("Pick-place", "Robot places cube inside the box (cube lands in box)", 0.5),
    ("Pick-place", "— MAX SCORE —", 1.0),
    (None, None, None),
    ("Stacking", "Robot grasps the red cube", 0.5),
    ("Stacking", "Robot places red cube on top of blue cube (stable)", 0.5),
    ("Stacking", "— MAX SCORE —", 1.0),
    (None, None, None),
    ("Sorting", "Robot grasps the red cube", 0.25),
    ("Sorting", "Robot places red cube in RIGHT box", 0.25),
    ("Sorting", "Robot grasps the blue cube", 0.25),
    ("Sorting", "Robot places blue cube in LEFT box", 0.25),
    ("Sorting", "— MAX SCORE —", 1.0),
    (None, None, None),
    ("NOTES", "Award partial credit — e.g. grasp succeeded but drop: 0.5", ""),
    ("NOTES", "Score 0 if robot fails to interact with any object", ""),
    ("NOTES", "Score 0.5 if grasp succeeds but placement fails", ""),
]

for r, (task, criterion, pts) in enumerate(guide_rows, start=1):
    if task == "TASK":
        _cell(ws_guide, r, 1, task,      fill=C_HEADER)
        _cell(ws_guide, r, 2, criterion, fill=C_HEADER)
        _cell(ws_guide, r, 3, pts,       fill=C_HEADER)
    elif task is None:
        ws_guide.row_dimensions[r].height = 6
    elif "MAX SCORE" in str(criterion):
        _cell(ws_guide, r, 1, task,      fill=C_SCORE, bold=True)
        _cell(ws_guide, r, 2, criterion, fill=C_SCORE, bold=True)
        _cell(ws_guide, r, 3, pts,       fill=C_SCORE, bold=True)
    elif task == "NOTES":
        _cell(ws_guide, r, 1, "NOTE",    fill="FFF2CC", bold=True)
        _cell(ws_guide, r, 2, criterion, fill="FFF2CC", align="left", wrap=True)
        _cell(ws_guide, r, 3, "",        fill="FFF2CC")
    else:
        _cell(ws_guide, r, 1, task,      fill=C_GUIDE, align="left")
        _cell(ws_guide, r, 2, criterion, fill=C_GUIDE, align="left", wrap=True)
        _cell(ws_guide, r, 3, pts,       fill=C_GUIDE)

ws_guide.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Pick-place Recording
# Sheet 3 — Stacking Recording
# ══════════════════════════════════════════════════════════════════════════════

TASKS = [
    ("Pick-place",  "Pick up the red cube and place it in the box."),
    ("Stacking",    "Put the red cube on top of the blue cube."),
]

CONDITIONS = [
    ("Baseline",   C_BASELINE),
    ("LP 4Hz",     C_LP),
    ("AJAS_Spline",C_AJAS),
]

POSITIONS    = [1, 2, 3, 4, 5]
CONFIGS      = ["A", "B"]
TRIALS       = [1, 2]

COLS = [
    "Condition", "Position", "Config", "Trial",
    "Score\n(0–1)",
    "Sm\n(auto)",
    "SER\n(auto)",
    "Notes",
]
COL_WIDTHS = [14, 10, 8, 7, 10, 10, 10, 30]

for task_name, task_desc in TASKS:
    ws = wb.create_sheet(title=task_name)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    title_cell = ws.cell(row=1, column=1,
                         value=f"{task_name} Evaluation — \"{task_desc}\"")
    title_cell.fill      = _fill(C_HEADER)
    title_cell.font      = Font(bold=True, color="FFFFFF", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border    = _border()
    ws.row_dimensions[1].height = 28

    # Column headers
    for c, (hdr, w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        _cell(ws, 2, c, hdr, fill=C_HEADER)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 30

    # Data rows
    row = 3
    for condition, cond_fill in CONDITIONS:
        for pos in POSITIONS:
            for cfg in CONFIGS:
                for trial in TRIALS:
                    _cell(ws, row, 1, condition, fill=cond_fill, align="left")
                    _cell(ws, row, 2, pos,       fill=cond_fill)
                    _cell(ws, row, 3, cfg,       fill=cond_fill)
                    _cell(ws, row, 4, trial,     fill=cond_fill)
                    _cell(ws, row, 5, "",        fill="FFFFFF")   # Score — manual entry
                    _cell(ws, row, 6, "",        fill="FFFFFF")   # Sm    — from CSV
                    _cell(ws, row, 7, "",        fill="FFFFFF")   # SER   — from CSV
                    _cell(ws, row, 8, "",        fill="FFFFFF")   # Notes
                    ws.row_dimensions[row].height = 16
                    row += 1

        # blank separator between conditions
        ws.row_dimensions[row].height = 6
        row += 1

    # Summary block at bottom
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    _cell(ws, row, 1, "SUMMARY", fill=C_HEADER)

    row += 1
    for condition, cond_fill in CONDITIONS:
        n_trials = len(POSITIONS) * len(CONFIGS) * len(TRIALS)
        _cell(ws, row, 1, condition,   fill=cond_fill, bold=True, align="left")
        _cell(ws, row, 2, "SR =",      fill=cond_fill)
        _cell(ws, row, 3, f"__ / {n_trials}", fill="FFFFFF")
        _cell(ws, row, 4, "Mean Sm =", fill=cond_fill)
        _cell(ws, row, 5, "",          fill="FFFFFF")
        _cell(ws, row, 6, "Mean SER =",fill=cond_fill)
        _cell(ws, row, 7, "",          fill="FFFFFF")
        _cell(ws, row, 8, "",          fill="FFFFFF")
        row += 1

    ws.freeze_panes = "A3"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Legend
# ══════════════════════════════════════════════════════════════════════════════
ws_leg = wb.create_sheet(title="Legend")
ws_leg.column_dimensions["A"].width = 18
ws_leg.column_dimensions["B"].width = 50

legend = [
    ("COLUMN",       "MEANING"),
    ("Condition",    "Baseline / LP 4Hz / AJAS_Spline — which processor was active"),
    ("Position",     "Cube start position (1–5, defined in your position grid)"),
    ("Config",       "A = orientation 1, B = orientation 2 (e.g. long-side vs short-side)"),
    ("Trial",        "1st or 2nd repeat of this position+config combo"),
    ("Score",        "Enter manually: 0 / 0.5 / 1.0 (or partial per scoring guide)"),
    ("Sm (auto)",    "Weighted mean frequency from eval script CSV — lower = smoother"),
    ("SER (auto)",   "Signal Energy Ratio — higher = more energy in low frequencies (smoother)"),
    ("Notes",        "Anything notable: grasped but dropped, near-miss, robot bumped table…"),
]

for r, (col, meaning) in enumerate(legend, start=1):
    fill = C_HEADER if r == 1 else C_GUIDE
    _cell(ws_leg, r, 1, col,     fill=fill, align="left")
    _cell(ws_leg, r, 2, meaning, fill=fill, align="left", wrap=True)
    ws_leg.row_dimensions[r].height = 18 if r > 1 else 22

ws_leg.freeze_panes = "A2"

# ── save ───────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved: {OUT}")
